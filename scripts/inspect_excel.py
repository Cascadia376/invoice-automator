import openpyxl
import os

file_path = r"c:\Users\Jay\Documents\Github\invoice-automator\design_files\Return Authorization Form Revised September 2022.xlsx"

try:
    wb = openpyxl.load_workbook(file_path)
    print(f"Sheet names: {wb.sheetnames}")
    
    sheet = wb.active
    print(f"Active sheet: {sheet.title}")
    
    print("\nFirst 20 rows:")
    for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
        # Filter out None values for cleaner output
        row_content = [str(cell) if cell is not None else "" for cell in row]
        # Only print non-empty rows
        if any(row_content):
            print(f"Row {i}: {row_content}")
        if i >= 30:
            break
            
except Exception as e:
    print(f"Error: {e}")
