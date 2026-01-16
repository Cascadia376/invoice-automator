import openpyxl

file_path = r"c:\Users\Jay\Documents\Github\invoice-automator\design_files\Return Authorization Form Revised September 2022.xlsx"

try:
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    print("--- Searching for 'Store Name' ---")
    found_store = False
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and "Store Name" in str(cell.value):
                print(f"Found 'Store Name' at {cell.coordinate}: {cell.value}")
                found_store = True
    
    print("\n--- Examining Block 1 (Rows 6-13) ---")
    for row in range(6, 14):
        vals = []
        for col in range(1, 8):
            cell = sheet.cell(row=row, column=col)
            vals.append(f"{cell.coordinate}={cell.value}")
        print(f"Row {row}: {', '.join(vals)}")

except Exception as e:
    print(f"Error: {e}")
