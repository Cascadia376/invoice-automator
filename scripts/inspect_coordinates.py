import openpyxl

file_path = r"c:\Users\Jay\Documents\Github\invoice-automator\design_files\Return Authorization Form Revised September 2022.xlsx"

try:
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    # Check Header Fields
    print("--- Header Area ---")
    for row in range(12, 20):
        for col in range(1, 10):
            cell = sheet.cell(row=row, column=col)
            if cell.value:
                print(f"({row}, {col}) {cell.coordinate}: {cell.value}")

    # Check Table Header
    print("\n--- Table Header Area ---")
    for row in range(20, 25):
        for col in range(1, 10):
            cell = sheet.cell(row=row, column=col)
            if cell.value:
                print(f"({row}, {col}) {cell.coordinate}: {cell.value}")
                
except Exception as e:
    print(f"Error: {e}")
