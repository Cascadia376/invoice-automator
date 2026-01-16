import sys
import os

# Add backend to path so we can import models/services
sys.path.append(os.path.join(os.getcwd(), "backend"))

from models import Invoice, LineItem
from services import ldb_service
import openpyxl

def test_ldb_generation():
    print("--- Testing LDB Generation ---")
    
    # Mock Data
    items = [
        LineItem(
            sku="12345", 
            description="Test Beer", 
            quantity=10, 
            issue_type="Breakage", 
            issue_notes="Bottle broken", 
            issue_status="open"
        ),
        LineItem(
            sku="67890", 
            description="Test Wine", 
            quantity=5, 
            issue_type="Missing", 
            issue_notes="Not in box",
            issue_status="open"
        ),
        LineItem(
            sku="11111", 
            description="Good Item", 
            quantity=50, 
            issue_type=None # Should be ignored
        )
    ]
    
    invoice = Invoice(
        invoice_number="INV-001",
        line_items=items,
        vendor_name="LDB"
    )
    
    print("Generating Excel...")
    try:
        excel_bytes = ldb_service.generate_ldb_return_form(invoice)
        print(f"Generated {len(excel_bytes)} bytes.")
        
        # Save to file for manual inspection if needed
        output_path = "test_ldb_output.xlsx"
        with open(output_path, "wb") as f:
            f.write(excel_bytes)
        print(f"Saved to {output_path}")
        
        # Verify Content with openpyxl
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        
        # Check Item 1 (Block 1 - SKU Row 8)
        # Reason: B7
        # SKU: B8
        val_reason = ws["B7"].value
        val_sku = ws["B8"].value
        print(f"Item 1 - Reason (B7): {val_reason} (Expected: Breakage)")
        print(f"Item 1 - SKU (B8): {val_sku} (Expected: 12345)")
        
        if val_sku != "12345":
            print("FAILED: SKU mismatch item 1")
            
        # Check Item 2 (Block 2 - SKU Row 16)
        # Reason: B15
        # SKU: B16
        val_reason_2 = ws["B15"].value
        val_sku_2 = ws["B16"].value
        print(f"Item 2 - Reason (B15): {val_reason_2} (Expected: Missing)")
        print(f"Item 2 - SKU (B16): {val_sku_2} (Expected: 67890)")
        
        if val_sku_2 != "67890":
            print("FAILED: SKU mismatch item 2")
            
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    test_ldb_generation()
