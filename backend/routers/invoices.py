from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, BackgroundTasks, Response
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
import shutil
import os
import uuid
import csv
import io
import fitz # PyMuPDF
import tempfile
import logging

logger = logging.getLogger(__name__)

import models, schemas, auth
from database import get_db
from services import parser, textract_service, vendor_service, product_service, storage, validation_service, export_service, ingestion_service, ldb_service, ldb_parser, splitting_service
from services.textract_service import parse_float

router = APIRouter(
    prefix="/api/invoices",
    tags=["invoices"]
)

# Use absolute path for uploads (mirroring main.py logic, though arguably should be in config)
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")

@router.post("/upload", response_model=List[schemas.Invoice])
async def upload_invoice(
    file: UploadFile = File(...), 
    db: Session = Depends(get_db),
    ctx: auth.UserContext = Depends(auth.get_current_user)
):
    if not ctx:
        print("UPLOAD FAIL: Missing user context")
        raise HTTPException(status_code=401, detail="Authentication required")
        
    print(f"UPLOAD REQUEST: User={ctx.user_id}, Org={ctx.org_id}, File={file.filename}")
    try:
        file_id = str(uuid.uuid4())
        file_ext = os.path.splitext(file.filename)[1].lower()
        temp_dir = tempfile.gettempdir()
        original_temp_path = os.path.join(temp_dir, f"original_{file_id}{file_ext}")
        
        with open(original_temp_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
            
        # 1. Decide if splitting is needed (only for PDFs)
        print("STAGE 1: Checking for multi-invoice content...")
        files_to_process = []
        if file_ext == ".pdf":
            print(f"DEBUG: Checking for multi-invoice content in {file.filename}")
            boundaries = splitting_service.detect_invoice_boundaries(original_temp_path)
            
            # Use splitting service to isolate invoices (removes cover pages even if only 1 invoice found)
            split_paths = splitting_service.split_pdf_into_files(original_temp_path, boundaries)
            files_to_process = split_paths
        else:
            files_to_process = [original_temp_path]

        # 2. Process each file through ingestion
        print(f"STAGE 2: Processing {len(files_to_process)} file(s)...")
        created_invoices = []
        for file_path in files_to_process:
            try:
                invoices = ingestion_service.process_invoice(
                    db=db,
                    file_path=file_path,
                    org_id=ctx.org_id,
                    user_id=ctx.user_id,
                    original_filename=file.filename
                )
                created_invoices.extend(invoices)
            except Exception as proc_error:
                print(f"ERROR processing file {file_path}: {proc_error}")
                import traceback
                traceback.print_exc()
                # Continue processing remaining files

        if not created_invoices:
            raise HTTPException(status_code=500, detail="Failed to process any invoices from the uploaded file")

        return created_invoices
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"ERROR processing invoice: {error_detail}")
        raise HTTPException(status_code=500, detail=f"Internal Server Error: {str(e)}\n{error_detail}")

from sqlalchemy import or_

# --- Stellar Routes (Must be before /{invoice_id}) ---

@router.post("/preflight-post", response_model=schemas.PreflightResponse)
def preflight_post_invoices(
    invoice_ids: List[str],
    db: Session = Depends(get_db),
    ctx: auth.UserContext = Depends(auth.get_current_user)
):
    """
    Check if invoices are ready for Stellar posting.
    Returns list of ready IDs, issues, and blocking vendor resolutions.
    """
    from services import stellar_service
    return stellar_service.check_invoice_preflight(db, invoice_ids)

@router.patch("/bulk-post")
async def bulk_post_invoices(
    invoice_ids: List[str],
    background_tasks: BackgroundTasks, # Use BG tasks for speed if many? No, user wants direct feedback.
    db: Session = Depends(get_db),
    ctx: auth.UserContext = Depends(auth.get_current_user)
):
    """
    Strict bulk post. 
    1. Runs preflight check internally (sanity check)
    2. Posts 'ready' invoices
    3. Returns results per invoice
    """
    from services import stellar_service
    
    # 1. Sanity Check
    preflight = stellar_service.check_invoice_preflight(db, invoice_ids)
    
    results = {
        "success": [],
        "failed": [],
        "skipped": [] # Issues/Blocking
    }
    
    # Process attributes from preflight
    # ready_ids are safe to post
    
    for inv_id in preflight["ready_ids"]:
        try:
             # Fetch invoice again to be safe/clean context
             db_inv = db.query(models.Invoice).filter(models.Invoice.id == inv_id).first()
             if not db_inv: 
                 continue
                 
             # Direct Post
             # We assume mapping exists because it passed preflight
             stellar_result = await stellar_service.post_invoice_if_configured(db_inv, db, require_config=True)
             
             db_inv.is_posted = True
             db_inv.status = 'posted'
             db.commit()
             
             results["success"].append({
                 "id": inv_id,
                 "asn": db_inv.stellar_asn_number
             })
             
        except Exception as e:
            logger.error(f"Bulk post failed for {inv_id}: {e}")
            results["failed"].append({
                "id": inv_id,
                "reason": str(e)
            })
            
    # Add skipped info
    for issue in preflight["issues"]:
        results["failed"].append({
            "id": issue["invoice_id"],
            "reason": f"{issue['issue_type']}: {issue['message']}"
        })
        
    for block in preflight["blocking_vendors"]:
        for inv_id in block["invoice_ids"]:
             results["failed"].append({
                "id": inv_id,
                "reason": f"Vendor '{block['vendor_name']}' unmapped"
            })

    return {"status": "completed", "results": results}



@router.get("", response_model=schemas.InvoiceListResponse)
def read_invoices(
    skip: int = 0, 
    limit: int = 100, 
    search: Optional[str] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    ctx: auth.UserContext = Depends(auth.get_current_user)
):
    print(f"FETCH REQUEST: User={ctx.user_id}, Org={ctx.org_id}, Skip={skip}, Limit={limit}, Search={search}, Status={status}")
    try:
        query = db.query(models.Invoice).filter(models.Invoice.organization_id == ctx.org_id)
        
        if search:
            query = query.filter(
                or_(
                    models.Invoice.invoice_number.ilike(f"%{search}%"),
                    models.Invoice.vendor_name.ilike(f"%{search}%")
                )
            )
        
        if status and status != 'all':
            # Handle custom 'issue' status which might be defined as invoices with issues
            if status == 'issue':
                query = query.filter(models.Invoice.line_items.any(models.LineItem.issue_type.isnot(None)))
            else:
                query = query.filter(models.Invoice.status == status)
                
        total = query.count()
        invoices = query.order_by(models.Invoice.created_at.desc()).offset(skip).limit(limit).all()
        
        # Point to proxy endpoint and add tenant for linking
        store = db.query(models.Store).filter(models.Store.organization_id == ctx.org_id).first()
        for inv in invoices:
            if inv.file_url:
                 inv.file_url = f"/api/invoices/{inv.id}/file"
            if store:
                inv.stellar_tenant = store.stellar_tenant
                 
        return {
            "items": invoices,
            "total": total,
            "skip": skip,
            "limit": limit
        }
    except Exception as e:
        print(f"ERROR fetching invoices: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/stats", response_model=schemas.DashboardStats)
def get_dashboard_stats(
    db: Session = Depends(get_db),
    ctx: auth.UserContext = Depends(auth.get_current_user)
):
    """Get global statistics for the dashboard"""
    base_query = db.query(models.Invoice).filter(models.Invoice.organization_id == ctx.org_id)
    
    total_invoices = base_query.count()
    needs_review = base_query.filter(models.Invoice.status == 'needs_review').count()
    approved = base_query.filter(or_(models.Invoice.status == 'approved', models.Invoice.status == 'pushed', models.Invoice.status == 'posted')).count()
    
    # Invoices with issues: Count active issues from the Issue model
    issue_count = db.query(models.Issue).filter(
        models.Issue.organization_id == ctx.org_id,
        or_(models.Issue.status == 'open', models.Issue.status == 'reported')
    ).count()
    
    # Calculate time saved (15 mins per approved invoice)
    hours_saved = (approved * 15) / 60
    time_saved_str = f"{hours_saved:.1f}h"
    
    return {
        "total_invoices": total_invoices,
        "needs_review": needs_review,
        "approved": approved,
        "issue_count": issue_count,
        "time_saved": time_saved_str
    }

@router.get("/{invoice_id}/file")
def get_invoice_file(
    invoice_id: str,
    db: Session = Depends(get_db),
    ctx: auth.UserContext = Depends(auth.get_current_user)
):
    """Proxy endpoint to stream files from S3 and avoid CORS issues"""
    invoice = db.query(models.Invoice).filter(
        models.Invoice.id == invoice_id,
        models.Invoice.organization_id == ctx.org_id
    ).first()
    
    if not invoice or not invoice.file_url:
        raise HTTPException(status_code=404, detail="File not found")
        
    s3_key = invoice.file_url
    # Validate key does not look like a proxy URL (DB corruption check)
    if s3_key.startswith("/api/"):
        logger.error(f"Invalid S3 Key in DB for invoice {invoice_id}: {s3_key}")
        raise HTTPException(status_code=500, detail="Invalid file reference in database")

    try:
        import boto3
        
        if not storage.AWS_BUCKET_NAME:
            logger.error("AWS_BUCKET_NAME not set in environment.")
            raise HTTPException(status_code=500, detail="Server storage parsing error (Missing Bucket Config)")
            
        s3 = storage.get_s3_client()
        response = s3.get_object(Bucket=storage.AWS_BUCKET_NAME, Key=s3_key)
        return StreamingResponse(
            response['Body'], 
            media_type=response.get('ContentType', 'application/pdf'),
            headers={
                "Content-Disposition": f"inline; filename=\"{invoice_id}.pdf\""
            }
        )
    except Exception as e:
        logger.error(f"PROXY ERROR: {e}")
        # Distinguish between not found and other errors if possible, 
        # but safely returning the error string helps debugging.
        if "NoSuchKey" in str(e) or "404" in str(e):
             raise HTTPException(status_code=404, detail=f"File not found in S3: {s3_key}")
        raise HTTPException(status_code=500, detail=f"Storage Error: {str(e)}")

from sqlalchemy.orm import joinedload

@router.get("/{invoice_id}", response_model=schemas.Invoice)
def read_invoice(
    invoice_id: str, 
    db: Session = Depends(get_db),
    ctx: auth.UserContext = Depends(auth.get_current_user)
):
    print(f"READ REQUEST: Fetching invoice {invoice_id}")
    invoice = db.query(models.Invoice).options(
        joinedload(models.Invoice.line_items),
        joinedload(models.Invoice.issues).joinedload(models.Issue.line_items)
    ).filter(
        models.Invoice.id == invoice_id,
        models.Invoice.organization_id == ctx.org_id
    ).first()
    
    if invoice is None:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    # Point to proxy endpoint
    if invoice.file_url:
         invoice.file_url = f"/api/invoices/{invoice.id}/file"

    # Fetch store/tenant for linking
    store = db.query(models.Store).filter(models.Store.organization_id == ctx.org_id).first()
    if store:
        invoice.stellar_tenant = store.stellar_tenant

    # Calculate Category Summary
    summary = {}
    for item in invoice.line_items:
        cat = item.category_gl_code or "Uncategorized"
        summary[cat] = summary.get(cat, 0.0) + (item.amount or 0.0)
    
    invoice.category_summary = {k: round(v, 2) for k, v in summary.items()}
         
    return invoice

@router.put("/{invoice_id}", response_model=schemas.Invoice)
def update_invoice(
    invoice_id: str, 
    invoice: schemas.InvoiceUpdate, 
    db: Session = Depends(get_db),
    ctx: auth.UserContext = Depends(auth.get_current_user)
):
    db_invoice = db.query(models.Invoice).filter(
        models.Invoice.id == invoice_id,
        models.Invoice.organization_id == ctx.org_id
    ).first()
    if db_invoice is None:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    update_data = invoice.dict(exclude_unset=True)
    
    # Handle line items separately
    if "line_items" in update_data:
        # Clear existing
        db.query(models.LineItem).filter(models.LineItem.invoice_id == invoice_id).delete()
        # Add new
        for item in update_data["line_items"]:
            # Learn SKU -> Category mapping if both are present
            sku = item.get("sku")
            category_gl_code = item.get("category_gl_code")
            
            if sku and category_gl_code:
                try:
                    # Check if mapping exists
                    mapping = db.query(models.SKUCategoryMapping).filter(
                        models.SKUCategoryMapping.sku == sku,
                        models.SKUCategoryMapping.category_gl_code == category_gl_code,
                        models.SKUCategoryMapping.organization_id == ctx.org_id
                    ).first()
                    
                    if mapping:
                        mapping.usage_count += 1
                    else:
                        new_mapping = models.SKUCategoryMapping(
                            id=str(uuid.uuid4()),
                            organization_id=ctx.org_id,
                            sku=sku,
                            category_gl_code=category_gl_code,
                            usage_count=1
                        )
                        db.add(new_mapping)
                except Exception as e:
                    print(f"WARNING: Could not learn SKU mapping: {e}")
            
            db_item = models.LineItem(id=str(uuid.uuid4()), invoice_id=invoice_id, **item)
            db.add(db_item)
        del update_data["line_items"]

    for key, value in update_data.items():
        setattr(db_invoice, key, value)

    db.commit()
    db.refresh(db_invoice)
    
    if db_invoice.file_url:
         db_invoice.file_url = f"/api/invoices/{db_invoice.id}/file"
         
    return db_invoice

@router.delete("/{invoice_id}")
def delete_invoice(
    invoice_id: str, 
    db: Session = Depends(get_db),
    ctx: auth.UserContext = Depends(auth.get_current_user)
):
    db_invoice = db.query(models.Invoice).filter(
        models.Invoice.id == invoice_id,
        models.Invoice.organization_id == ctx.org_id
    ).first()
    if db_invoice is None:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    db.delete(db_invoice)
    db.commit()
    return {"status": "success", "message": "Invoice deleted"}

@router.patch("/{invoice_id}/post", response_model=schemas.Invoice)
async def post_invoice_to_pos(
    invoice_id: str,
    db: Session = Depends(get_db),
    ctx: auth.UserContext = Depends(auth.get_current_user)
):
    """
    Post an invoice to POS system (Stellar).
    
    This endpoint:
    1. Marks the invoice as posted in our system
    2. Attempts to post to Stellar if configured for the vendor
    3. Returns the updated invoice with Stellar posting status
    """
    from services import stellar_service
    
    db_invoice = db.query(models.Invoice).filter(
        models.Invoice.id == invoice_id,
        models.Invoice.organization_id == ctx.org_id
    ).first()
    if db_invoice is None:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    # Attempt to post to Stellar
    # We use strict mode (require_config=True) because this is a user-initiated action
    try:
        # Auto-map vendor if not configured
        # Use robust lookup (handles aliases like "Inc", "Ltd")
        vendor = None
        if db_invoice.vendor_id:
            vendor = db.query(models.Vendor).filter(models.Vendor.id == db_invoice.vendor_id).first()
        elif db_invoice.vendor_name:
            vendor = vendor_service.find_vendor_by_name(db, db_invoice.vendor_name, ctx.org_id)
            
        if vendor:
            # If we found the vendor object, ensure it's mapped to Stellar
            # This handles cases where vendor exists in DB but has no Stellar ID
            await stellar_service.ensure_vendor_mapping(db, vendor)
        else:
            # If we didn't find the vendor in our DB, we might want to auto-create it?
            # For now, just logging. post_invoice_if_configured will likely fail.
            logger.warning(f"Could not resolve vendor '{db_invoice.vendor_name}' for auto-mapping.")

        stellar_result = await stellar_service.post_invoice_if_configured(db_invoice, db, require_config=True)
        
        # If successful (or if logic allowed skipping without error, which shouldn't happen with require_config=True)
        db_invoice.is_posted = True
        db_invoice.status = 'posted'
        logger.info(f"Successfully posted invoice {invoice_id} to Stellar")
        
    except stellar_service.StellarError as e:
        # If posting fails, we do NOT celebrate.
        # We ensure is_posted is False (just in case)
        db_invoice.is_posted = False
        # And we re-raise as a 400 bad request so the UI knows it failed
        logger.error(f"Failed to post invoice {invoice_id} to Stellar: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))
        
    except Exception as e:
        logger.exception(f"Unexpected error posting invoice {invoice_id} to Stellar")
        raise HTTPException(status_code=500, detail=f"Unexpected error: {str(e)}")
    
    db.commit()
    db.refresh(db_invoice)
    
    # Normalize file URL for response
    if db_invoice.file_url:
         db_invoice.file_url = f"/api/invoices/{db_invoice.id}/file"
    
    # Add tenant for linking
    store = db.query(models.Store).filter(models.Store.organization_id == ctx.org_id).first()
    if store:
        db_invoice.stellar_tenant = store.stellar_tenant
    
    return db_invoice



@router.get("/stats/category-summary")
def get_category_summary(
    month: Optional[str] = None, # YYYY-MM
    db: Session = Depends(get_db),
    ctx: auth.UserContext = Depends(auth.get_current_user)
):
    query = db.query(models.Invoice).filter(
        models.Invoice.organization_id == ctx.org_id,
        or_(models.Invoice.status == 'approved', models.Invoice.status == 'posted'),
        models.Invoice.is_posted == True
    )
    
    if month:
        # Simple string-based filter for date (assuming YYYY-MM-DD format in DB)
        query = query.filter(models.Invoice.date.like(f"{month}%"))
        
    invoices = query.all()
    
    summary = {}
    total_tax = 0.0
    total_deposit = 0.0
    total_amount = 0.0
    
    for inv in invoices:
        total_tax += (inv.tax_amount or 0.0)
        total_deposit += (inv.deposit_amount or 0.0)
        total_amount += (inv.total_amount or 0.0)
        
        for item in inv.line_items:
            cat = item.category_gl_code or "Uncategorized"
            summary[cat] = summary.get(cat, 0.0) + (item.amount or 0.0)
            
    return {
        "category_totals": {k: round(v, 2) for k, v in summary.items()},
        "total_tax": round(total_tax, 2),
        "total_deposit": round(total_deposit, 2),
        "total_amount": round(total_amount, 2),
        "invoice_count": len(invoices)
    }

@router.post("/{invoice_id}/feedback")
def submit_feedback(
    invoice_id: str, 
    feedback_data: schemas.InvoiceUpdate, 
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    ctx: auth.UserContext = Depends(auth.get_current_user)
):
    db_invoice = db.query(models.Invoice).filter(
        models.Invoice.id == invoice_id,
        models.Invoice.organization_id == ctx.org_id
    ).first()
    if db_invoice is None:
        raise HTTPException(status_code=404, detail="Invoice not found")

    s3_key = db_invoice.file_url
    
    # 1. Trigger legacy template refinement in background
    if s3_key and not s3_key.startswith("http"):
        temp_dir = tempfile.gettempdir()
        temp_file_path = os.path.join(temp_dir, f"feedback_{invoice_id}.pdf")
        storage.download_file(s3_key, temp_file_path)
        
        background_tasks.add_task(
            parser.refine_template_with_feedback, 
            temp_file_path, 
            feedback_data.dict(exclude_unset=True),
            ctx.org_id
        )
    else:
        # Fallback for old local files (if any exist)
        relative_path = db_invoice.file_url.lstrip("/")
        file_path = os.path.join(BASE_DIR, relative_path)
        if os.path.exists(file_path):
            background_tasks.add_task(
                parser.refine_template_with_feedback, 
                file_path, 
                feedback_data.dict(exclude_unset=True),
                ctx.org_id
            )

    # 2. Trigger new vendor correction learning
    if db_invoice.vendor_id:
        feedback_dict = feedback_data.dict(exclude_unset=True)
        # Check standard fields
        fields_to_check = ['total_amount', 'subtotal', 'tax_amount', 'deposit_amount', 'shipping_amount', 'date', 'invoice_number']
        for field in fields_to_check:
            if field in feedback_dict:
                new_val = str(feedback_dict[field])
                old_val = str(getattr(db_invoice, field)) if getattr(db_invoice, field) is not None else ""
                
                if new_val != old_val:
                    vendor_service.learn_from_correction(
                        db,
                        invoice_id,
                        db_invoice.vendor_id,
                        ctx.org_id,
                        field,
                        old_val,
                        new_val,
                        raw_extraction_results=db_invoice.raw_extraction_results,
                        user_id=ctx.user_id
                    )

    return {"status": "success", "message": "Feedback received, refining template in background"}

@router.get("/{invoice_id}/validate")
def validate_invoice_endpoint(
    invoice_id: str,
    db: Session = Depends(get_db),
    ctx: auth.UserContext = Depends(auth.get_current_user)
):
    invoice = db.query(models.Invoice).filter(
        models.Invoice.id == invoice_id,
        models.Invoice.organization_id == ctx.org_id
    ).first()
    
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
        
    return validation_service.validate_invoice(db, invoice)

@router.get("/{invoice_id}/highlights")
def get_invoice_highlights(
    invoice_id: str,
    db: Session = Depends(get_db),
    ctx: auth.UserContext = Depends(auth.get_current_user)
):
    invoice = db.query(models.Invoice).filter(
        models.Invoice.id == invoice_id,
        models.Invoice.organization_id == ctx.org_id
    ).first()
    
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")

    s3_key = invoice.file_url
    temp_dir = tempfile.gettempdir()
    temp_file_path = os.path.join(temp_dir, f"highlights_{invoice_id}.pdf")
    
    try:
        if s3_key and not s3_key.startswith("http") and not os.path.exists(s3_key):
             storage.download_file(s3_key, temp_file_path)
        elif s3_key and os.path.exists(s3_key):
             shutil.copy(s3_key, temp_file_path)
        elif invoice.file_url and invoice.file_url.startswith("/"):
             local_path = os.path.join(BASE_DIR, invoice.file_url.lstrip("/"))
             if os.path.exists(local_path):
                shutil.copy(local_path, temp_file_path)
             else:
                return {} 
        else:
             return {}
             
        # Search for highlights
        highlights = {}
        doc = fitz.open(temp_file_path)
        
        def add_highlight(field_name, text_value):
            if not text_value or str(text_value).lower() == "unknown":
                return
            
            text_str = str(text_value)
            clean_text = text_str.replace("$", "").replace(",", "").strip()
            
            found = []
            for page_num, page in enumerate(doc):
                rects = page.search_for(text_str)
                if not rects and clean_text != text_str:
                    rects = page.search_for(clean_text)
                
                for r in rects:
                    width = page.rect.width
                    height = page.rect.height
                    
                    found.append({
                        "page": page_num + 1,
                        "rect": [r.x0, r.y0, r.x1 - r.x0, r.y1 - r.y0],
                        "norm": [r.x0/width, r.y0/height, (r.x1-r.x0)/width, (r.y1-r.y0)/height]
                    })
            
            if found:
                highlights[field_name] = found

        add_highlight("invoice_number", invoice.invoice_number)
        add_highlight("total_amount", f"{invoice.total_amount:.2f}")
        add_highlight("total_amount", f"{invoice.total_amount:,.2f}")
        add_highlight("subtotal", f"{invoice.subtotal:.2f}")
        add_highlight("subtotal", f"{invoice.subtotal:,.2f}")
        add_highlight("tax_amount", f"{invoice.tax_amount:.2f}")
        add_highlight("tax_amount", f"{invoice.tax_amount:,.2f}")
        add_highlight("deposit_amount", f"{invoice.deposit_amount:.2f}")
        add_highlight("deposit_amount", f"{invoice.deposit_amount:,.2f}")
        add_highlight("shipping_amount", f"{invoice.shipping_amount:.2f}")
        add_highlight("shipping_amount", f"{invoice.shipping_amount:,.2f}")
        add_highlight("discount_amount", f"{invoice.discount_amount:.2f}")
        add_highlight("discount_amount", f"{invoice.discount_amount:,.2f}")
        add_highlight("po_number", invoice.po_number)
        add_highlight("date", invoice.date)
        add_highlight("vendor_name", invoice.vendor_name)
        
        for idx, item in enumerate(invoice.line_items):
            prefix = f"line_items[{idx}]"
            add_highlight(f"{prefix}.amount", f"{item.amount:.2f}")
            add_highlight(f"{prefix}.unit_cost", f"{item.unit_cost:.2f}")
            add_highlight(f"{prefix}.quantity", str(int(item.quantity) if item.quantity.is_integer() else item.quantity))
            add_highlight(f"{prefix}.sku", item.sku)
            add_highlight(f"{prefix}.description", item.description)
            add_highlight(f"{prefix}.units_per_case", str(int(item.units_per_case) if item.units_per_case.is_integer() else item.units_per_case))
            add_highlight(f"{prefix}.cases", str(int(item.cases) if item.cases.is_integer() else item.cases))

        doc.close()
        if os.path.exists(temp_file_path):
            os.remove(temp_file_path)
            
        return highlights

    except Exception as e:
        print(f"Error generating highlights: {e}")
        return {}

@router.get("/{invoice_id}/validate")
def validate_invoice(
    invoice_id: str,
    db: Session = Depends(get_db),
    ctx: auth.UserContext = Depends(auth.get_current_user)
):
    invoice = db.query(models.Invoice).filter(
        models.Invoice.id == invoice_id,
        models.Invoice.organization_id == ctx.org_id
    ).first()
    
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
        
    line_item_warnings = {}
    global_warnings = []
    
    # 1. Math Check (Quantity * Unit Cost = Amount)
    for item in invoice.line_items:
        if abs((item.quantity * item.unit_cost) - item.amount) > 0.02:
            warnings = line_item_warnings.get(item.id, [])
            warnings.append(f"Math Error: {item.quantity} * {item.unit_cost} = {item.quantity * item.unit_cost:.2f} (Invoice says {item.amount:.2f})")
            line_item_warnings[item.id] = warnings

    # 2. Product Master Data Check (via Supabase/Cache)
    for item in invoice.line_items:
        validation = product_service.validate_item_against_master(db, ctx.org_id, {
            "sku": item.sku,
            "description": item.description,
            "units_per_case": item.units_per_case,
            "quantity": item.quantity,
            "unit_cost": item.unit_cost,
            "amount": item.amount
        })
        
        if validation["status"] == "success" and validation["flags"]:
            warnings = line_item_warnings.get(item.id, [])
            warnings.extend(validation["flags"])
            line_item_warnings[item.id] = warnings

    # 3. Global Checks (Sum of line items vs subtotal)
    sum_items = sum(item.amount for item in invoice.line_items)
    if abs(sum_items - (invoice.subtotal or 0)) > 0.05:
        global_warnings.append(f"Subtotal Mismatch: Sum of items ({sum_items:.2f}) ≠ Subtotal ({invoice.subtotal:.2f})")

    return {
        "global_warnings": global_warnings,
        "line_item_warnings": line_item_warnings
    }

@router.get("/{invoice_id}/export/csv")
def export_invoice_csv(
    invoice_id: str, 
    columns: str = None, 
    db: Session = Depends(get_db),
    ctx: auth.UserContext = Depends(auth.get_current_user)
):
    invoice = db.query(models.Invoice).filter(
        models.Invoice.id == invoice_id,
        models.Invoice.organization_id == ctx.org_id
    ).first()
    
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    csv_content = export_service.generate_csv(invoice)
    
    headers = {
        'Content-Disposition': f'attachment; filename="invoice_{invoice.invoice_number or invoice_id}.csv"'
    }
    
    return StreamingResponse(iter([csv_content]), media_type="text/csv", headers=headers)

@router.get("/{invoice_id}/export/excel")
def export_invoice_excel(
    invoice_id: str, 
    db: Session = Depends(get_db),
    ctx: auth.UserContext = Depends(auth.get_current_user)
):
    invoice = db.query(models.Invoice).filter(
        models.Invoice.id == invoice_id,
        models.Invoice.organization_id == ctx.org_id
    ).first()
    
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from datetime import datetime
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice Export"
    
    # Headers based on user request: SKU, Receiving Qty (UOM), Confirmed total
    headers = ["SKU", "Receiving Qty (UOM)", "Confirmed total"]
    ws.append(headers)
    
    # Make headers bold
    for cell in ws[1]:
        cell.font = Font(bold=True)
        
    for item in invoice.line_items:
        ws.append([
            item.sku or "N/A",
            item.quantity,
            item.amount
        ])
        
    # Filename: [Supplier Name] - [Date] - [PO Number].xlsx
    safe_vendor = "".join(x for x in (invoice.vendor_name or "Unknown") if x.isalnum() or x in " -_").strip()
    safe_date = invoice.date or datetime.now().strftime("%Y-%m-%d")
    safe_po = invoice.po_number or "NO-PO"
    filename = f"{safe_vendor} - {safe_date} - {safe_po}.xlsx"
    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=\"{filename}\"",
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
    )

@router.get("/{invoice_id}/export/ldb")
def export_invoice_ldb_report(
    invoice_id: str, 
    db: Session = Depends(get_db),
    ctx: auth.UserContext = Depends(auth.get_current_user)
):
    invoice = db.query(models.Invoice).filter(
        models.Invoice.id == invoice_id,
        models.Invoice.organization_id == ctx.org_id
    ).first()
    
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    # 1. Validation: Ensure it is an LDB invoice
    vendor_name = (invoice.vendor_name or "").lower()
    valid_names = ["ldb", "liquor distribution branch", "bc liquor"]
    is_ldb = any(name in vendor_name for name in valid_names)
    
    if not is_ldb:
        raise HTTPException(status_code=400, detail="LDB Issue Reports can only be generated for LDB invoices.")

    # 2. Generate Report
    excel_content = ldb_service.generate_ldb_return_form(invoice)
    
    # 3. Persist to Storage
    # Filename: LDB_Issue_Report_[Invoice#]_[Date].xlsx
    safe_invoice = "".join(x for x in (invoice.invoice_number or "Unknown") if x.isalnum() or x in "-_").strip()
    safe_date = invoice.date or datetime.now().strftime("%Y-%m-%d")
    filename = f"LDB_Issue_Report_{safe_invoice}_{safe_date}.xlsx"
    
    # Save temp
    temp_path = f"/tmp/{filename}"
    with open(temp_path, "wb") as f:
        f.write(excel_content)
        
    # Upload S3
    s3_key = f"invoices/{ctx.org_id}/reports/{filename}"
    if storage.upload_file(temp_path, s3_key):
        # Update DB with link
        invoice.ldb_report_url = s3_key
        db.commit()
    
    # Clean up
    if os.path.exists(temp_path):
        os.remove(temp_path)

    return StreamingResponse(
        io.BytesIO(excel_content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=\"{filename}\"",
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
    )

@router.post("/export/excel/bulk")
def export_invoices_bulk(
    invoice_ids: List[str],
    db: Session = Depends(get_db),
    ctx: auth.UserContext = Depends(auth.get_current_user)
):
    invoices = db.query(models.Invoice).filter(
        models.Invoice.id.in_(invoice_ids),
        models.Invoice.organization_id == ctx.org_id
    ).all()
    
    if not invoices:
        raise HTTPException(status_code=404, detail="No invoices found")
        
    import zipfile
    from openpyxl import Workbook
    from datetime import datetime
    
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zip_file:
        for invoice in invoices:
            wb = Workbook()
            ws = wb.active
            headers = ["SKU", "Receiving Qty (UOM)", "Confirmed total"]
            ws.append(headers)
            
            for item in invoice.line_items:
                ws.append([item.sku or "N/A", item.quantity, item.amount])
                
            safe_vendor = "".join(x for x in (invoice.vendor_name or "Unknown") if x.isalnum() or x in " -_").strip()
            safe_date = invoice.date or datetime.now().strftime("%Y-%m-%d")
            safe_po = invoice.po_number or "NO-PO"
            filename = f"{safe_vendor} - {safe_date} - {safe_po}.xlsx"
            
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            zip_file.writestr(filename, excel_buffer.getvalue())
            
    zip_buffer.seek(0)
    
    return Response(
        content=zip_buffer.getvalue(),
        media_type="application/zip",
        headers={
            "Content-Disposition": f"attachment; filename=\"Invoices_Export_{datetime.now().strftime('%Y%m%d')}.zip\"",
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
    )

@router.post("/export/excel/bulk-approved")
def export_invoices_bulk_approved(
    invoice_ids: List[str],
    db: Session = Depends(get_db),
    ctx: auth.UserContext = Depends(auth.get_current_user)
):
    """
    Export specific invoices as individual XLSX files in a ZIP.
    Strictly filters for APPROVED invoices only.
    Items must be passed as IDs (usually from the current view).
    """
    # 1. Fetch Invoices from ID list
    invoices = db.query(models.Invoice).filter(
        models.Invoice.id.in_(invoice_ids),
        models.Invoice.organization_id == ctx.org_id
    ).all()
    
    if not invoices:
        raise HTTPException(status_code=404, detail="No invoices found")
        
    # 2. Filter for APPROVED only
    approved_invoices = [inv for inv in invoices if inv.status == 'approved']
    
    if not approved_invoices:
        raise HTTPException(status_code=400, detail="None of the selected invoices are approved.")
        
    import zipfile
    from datetime import datetime
    
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zip_file:
        for invoice in approved_invoices:
            # Generate XLSX content
            xlsx_bytes = export_service.generate_invoice_xlsx(invoice)
            
            # Filename: SupplierName - InvoiceNumber - Date.xlsx
            safe_vendor = "".join(x for x in (invoice.vendor_name or "Unknown") if x.isalnum() or x in " -_").strip()
            safe_invoice_num = "".join(x for x in (invoice.invoice_number or "NO-NUM") if x.isalnum() or x in "-_").strip()
            safe_date = invoice.date or datetime.now().strftime("%Y-%m-%d")
            
            filename = f"{safe_vendor} - {safe_invoice_num} - {safe_date}.xlsx"
            
            zip_file.writestr(filename, xlsx_bytes)
            
    zip_buffer.seek(0)
    
    return Response(
        content=zip_buffer.getvalue(),
        media_type="application/zip",
        headers={
            "Content-Disposition": f"attachment; filename=\"Approved_Invoices_{datetime.now().strftime('%Y%m%d')}.zip\"",
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
    )
