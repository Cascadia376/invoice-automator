import os
import io
from openpyxl import load_workbook
from models import Invoice

def generate_ldb_return_form(invoice: Invoice) -> bytes:
    """
    Generates an LDB Return Authorization Form (Excel) for items with issues.
    Populates 'Return Blocks' in the template.
    """
    # Path to the design file - assuming running from root or backend
    # Try different common paths to be robust
    possible_paths = [
        "design_files/Return Authorization Form Revised September 2022.xlsx",
        "../design_files/Return Authorization Form Revised September 2022.xlsx",
        r"c:\Users\Jay\Documents\Github\invoice-automator\design_files\Return Authorization Form Revised September 2022.xlsx"
    ]
    
    template_path = None
    for path in possible_paths:
        if os.path.exists(path):
            template_path = path
            break
            
    if not template_path:
        raise FileNotFoundError("LDB Excel Template not found.")

    wb = load_workbook(template_path)
    ws = wb.active

    # --- Header Info (Best Effort) ---
    # Attempt to fill Date if we can find a reasonable place, 
    # but based on analysis, strict coordinates were not definitive for headers 
    # except maybe finding "Date:" or "Store Name" labels. 
    # For now, we will focus on the line items as per the plan.

    # --- Line Item Population ---
    # Block 1 starts with SKU at Row 8.
    # Stride is 8 rows.
    # We can fill up to 5 blocks (Row 8, 16, 24, 32, 40).
    
    start_row = 8
    stride = 8
    max_blocks = 5
    
    # NEW: Fetch from the Issue table
    for i, issue in enumerate(invoice.issues):
        if i >= max_blocks:
            break # Template only holds 5 items per page
            
        current_sku_row = start_row + (i * stride)
        
        # We assume for LDB that an issue links to at least one line item
        # If multiple items are linked, we'll take the first for the SKU/Description
        item = issue.line_items[0] if issue.line_items else None
        
        # Reason: B{R-1}
        ws.cell(row=current_sku_row - 1, column=2, value=issue.type.replace('_', ' ').capitalize() if issue.type else "Generic Issue")
        
        if item:
            # SKU: B{R}
            ws.cell(row=current_sku_row, column=2, value=item.sku or "UNKNOWN")
            
            # Product Name: B{R+1}
            ws.cell(row=current_sku_row + 1, column=2, value=item.description or "")
            
            # Quantity: B{R+2}
            ws.cell(row=current_sku_row + 2, column=2, value=item.quantity)
        else:
            ws.cell(row=current_sku_row, column=2, value="N/A")
            ws.cell(row=current_sku_row + 1, column=2, value=issue.description or "")
            ws.cell(row=current_sku_row + 2, column=2, value=1)
        
        # Invoice #: B{R+3}
        ws.cell(row=current_sku_row + 3, column=2, value=invoice.invoice_number or "")
        
        # Comments: B{R+4} (Notes)
        # Pull latest communication note if available, otherwise description
        notes = issue.description or ""
        if issue.communications:
            latest_note = next((c.content for c in reversed(issue.communications) if c.type == 'note'), None)
            if latest_note:
                notes = latest_note
                
        ws.cell(row=current_sku_row + 4, column=2, value=notes)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
