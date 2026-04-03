import csv
import io
from typing import Any
from models import Invoice


def format_receiving_quantity(item: Any) -> float | int:
    """Return the receiving quantity in cases when available."""
    cases = getattr(item, "cases", None)
    try:
        if cases not in (None, "") and float(cases) > 0:
            value = float(cases)
        else:
            quantity = float(getattr(item, "quantity", 0) or 0)
            units_per_case = float(getattr(item, "units_per_case", 0) or 0)
            if quantity > 0 and units_per_case > 0:
                value = quantity / units_per_case
            else:
                value = quantity
    except (TypeError, ValueError):
        value = float(getattr(item, "quantity", 0) or 0)

    return int(value) if float(value).is_integer() else round(value, 2)


def generate_csv(invoice: Invoice) -> str:
    """
    Generates a CSV string for the given invoice.
    Format: Standard Import (SKU, Cases, Cost, Total)
    """
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Headers
    writer.writerow([
        "SKU", 
        "Receiving Qty (UOM)", 
        "Confirmed total"
    ])
    
    for item in invoice.line_items:
        writer.writerow([
            item.sku or "",
            format_receiving_quantity(item),
            f"{item.amount:.2f}" if item.amount is not None else ""
        ])
        
    return output.getvalue()

def generate_ldb_report(invoice: Invoice) -> bytes:
    """
    Generates an LDB Issue Report (Excel) for items with issues.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    
    wb = Workbook()
    ws = wb.active
    ws.title = "LDB Issues"
    
    headers = ["Invoice #", "Date", "SKU", "Description", "Issue Type", "Status", "Notes", "Qty", "Amount"]
    ws.append(headers)
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="36454F", end_color="36454F", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        
    has_issues = False
    for item in invoice.line_items:
        if item.issue_type: # Only include items with issues
            has_issues = True
            ws.append([
                invoice.invoice_number or "UNKNOWN",
                invoice.date or "",
                item.sku or "",
                item.description or "",
                item.issue_type,
                item.issue_status or "open",
                item.issue_notes or item.issue_description or "",
                item.quantity,
                item.amount
            ])
            
    if not has_issues:
        ws.append(["No issues flagged on this invoice."])
        
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
        

def generate_invoice_xlsx(invoice: Invoice) -> bytes:
    """
    Generates a single XLSX file for the given invoice.
    Format: SKU, Receiving Qty (Cases), Confirmed Total Cost
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice Export"
    
    # Headers
    headers = ["SKU", "Receiving Qty (UOM)", "Confirmed Total Cost"]
    ws.append(headers)
    
    # Bold Headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        
    for item in invoice.line_items:
        ws.append([
            item.sku or "",
            format_receiving_quantity(item),
            item.amount # Confirmed Total Cost
        ])
        
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
        
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
